

from scrapper.dian_web import DianRutSearcher
from use_cases.clasify_rut import clasify_rut
from use_cases.export import export_to_csv
import traceback
# Abre el archivo Excel
import openpyxl

from use_cases.process_names import normalize, unify_names 

file_name = 'VALIDACIÓN TERCEROS DIAN TATIS.xlsx'
wb = openpyxl.load_workbook(file_name)

ws = wb.active

print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
rut_searcher = DianRutSearcher()
ruts = []

def process_rut(rut_number, ruts):

  # Si no es valido el rut debe quedar en blanco el nombre y todos los campos menos RUT
  rut_searcher.search_rut(rut_number)
  rut_info = rut_searcher.get_rut_info()
  clasification = clasify_rut(rut_info)
  if rut_info:
    rut_info['clasificación'] = clasification
  rut_info = unify_names(rut_info)
  rut_info['RUT'] = rut_number
  rut_info = normalize(rut_info)
  ruts.append(rut_info)
  rut_searcher.reset_search()


try:
  for row_number in range(286, ws.max_row):
    print('Analyzing row number: ' + str(row_number) + '.')
    rut = ws.cell(row=row_number, column=2).value
    process_rut(rut, ruts)
    print('RUT analyzed: ' + str(rut) + '.')
    print(rut)
    if row_number > 1000:
      break
except Exception as e:
  print(e)
  traceback.print_exc()
finally:
  export_to_csv(ruts, 'rut1')

# rut_numbers = ['1071632198', '1019130051']




# export_to_csv(ruts, 'rut')
